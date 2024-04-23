from PIL import Image
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import time

def carregar_imagem(caminho):
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo '{caminho}' não encontrado.")
    
    try:
        imagem = Image.open(caminho)
        return imagem
    except Exception as e:
        print("Ocorreu um erro ao carregar a imagem:", e)

def imagem_para_matriz(imagem):
    largura, altura = imagem.size
    matriz = []

    for y in range(altura):
        linha = []
        for x in range(largura):
            pixel = imagem.getpixel((x, y))
            linha.append(pixel)
        matriz.append(linha)

    return matriz

def matriz_para_excel(matriz, caminho_imagem):
    wb = Workbook()
    ws = wb.active

    for y, linha in enumerate(matriz, start=1):
        for x, pixel in enumerate(linha, start=1):
            valor_pixel = f"({pixel[0]}, {pixel[1]}, {pixel[2]})"
            cell = ws.cell(row=y, column=x, value=valor_pixel)
            cell.alignment = cell.alignment.copy(wrapText=True)

    for coluna in ws.columns:
        max_length = 0
        coluna_letra = coluna[0].column_letter
        for cell in coluna:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[coluna_letra].width = adjusted_width

    nome_arquivo_original = os.path.splitext(os.path.basename(caminho_imagem))[0]
    nome_arquivo_convertido = nome_arquivo_original + '_converted.xlsx'
    caminho_arquivo = os.path.join(os.path.dirname(__file__), nome_arquivo_convertido)
    wb.save(caminho_arquivo)

    print(f"Matriz salva com sucesso no arquivo: {caminho_arquivo}")

def main():
    print("Imagens .png & .jpg são aceitas.")
    time.sleep(1.5)
    nome_arquivo = input("Digite o nome do arquivo de imagem (sem o formato): ")
    
    caminho_png = nome_arquivo + ".png"
    caminho_jpg = nome_arquivo + ".jpg"

    if os.path.exists(caminho_png):
        caminho_imagem = caminho_png
    elif os.path.exists(caminho_jpg):
        caminho_imagem = caminho_jpg
    else:
        print("Arquivo não encontrado.")
        return
    
    imagem = carregar_imagem(caminho_imagem)
    
    if imagem:
        print("Imagem carregada com sucesso!")
        matriz_pixels = imagem_para_matriz(imagem)
        print("Matriz de pixels criada.")
        print("Dimensões da matriz:", len(matriz_pixels), "x", len(matriz_pixels[0]))

        matriz_para_excel(matriz_pixels, caminho_imagem)


if __name__ == "__main__":
    main()
